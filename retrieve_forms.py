import requests
from requests.auth import HTTPBasicAuth
import pandas as pd
import logging
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from auth_vars import usrnm as a, passw as b, frmids as f

# Retrieve credentials and form IDs from environment variables
username = a
password = b
form_ids = f

# API base URL
base_api_url = 'https://intralog.com.ar/wp-json/custom/v1/form-submissions/'

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create an empty DataFrame with the specified columns
columns = ["Empresa", "Fecha creacion", "Razon social", "Nombre y apellido", "Servicio", "Mail", "Telefono", "Origen", "Sub-origen", "Mensaje", "Avance", "Estado", "form_id", "submission id"]
df_template = pd.DataFrame(columns=columns)

def get_form_submissions(api_url, username, password):
    try:
        response = requests.get(api_url, auth=HTTPBasicAuth(username, password))
        if response.status_code == 200:
            logger.info(f"Successfully retrieved data from {api_url}")
            return response.json()
        else:
            logger.error(f"Failed to retrieve data from {api_url}: {response.status_code}")
            logger.error(f"Error message: {response.text}")
            return None
    except requests.RequestException as e:
        logger.error(f"Request failed for {api_url}: {e}")
        return None

def process_submission(submission, form_id):
    servicio = submission.get('Me interesa el servicio' if form_id in [3, 4, 5] else 'Ubicación')
    avance, estado = determine_avance_estado(servicio, form_id)

    if form_id in [4, 5]:
        razon_social_key = "Razón Social"
        email_key = "Correo electrónico"
    else:
        if form_id == 3:
            email_key = "E-mail"
        else:
            email_key = "E-Mail"
        razon_social_key = "Razón social"
    
    processed = {
        "Empresa": "INTRALOG" if form_id in [3, 4, 5] else "INTRAPAL",
        "submission id": submission.get('id'),
        "Fecha creacion": submission.get('created_at'),
        # "Fecha actualizacion": submission.get('updated_at'),
        "Razon social": submission.get(razon_social_key),
        "Nombre y apellido": submission.get('Razón social' if form_id == 3 else 'Nombre y Apellido'),
        "Telefono": submission.get('Teléfono' if form_id != 3 else 'Telefono'),
        "Mail": submission.get(email_key),
        "Mensaje": submission.get('Mensaje' if form_id in [3, 4, 5] else 'Mensaje'),
        "Servicio": servicio,
        "Origen" : "Web",
        "Sub-origen" : "Signos",
        "Avance": avance,
        "Estado": estado,
        "form_id": form_id
    }
    return processed

def determine_avance_estado(servicio, form_id):
    if form_id != 7:
        if servicio == "Busco trabajo/ Ofrezco productos o servicios":
            return "▓", "negativo"
        else:
            return "⊕", "esperando datos"
    else:
        return "⊕", "esperando datos"

def save_to_excel(data, output_file):
    if not os.path.exists(output_file):
        # Save new data to a new workbook
        data.to_excel(output_file, index=False, engine='openpyxl')
        logger.info(f"DataFrame saved as new file {output_file}")
    else:
        # Load the existing workbook
        book = load_workbook(output_file)
        sheet = book.active
        
        # Load existing data into a DataFrame
        data_rows = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True))
        headers = [cell.value for cell in sheet[1]]
        existing_df = pd.DataFrame(data_rows, columns=headers)
        
        # Combine new data with existing data, avoiding duplicates
        updated_df = pd.concat([existing_df, data]).drop_duplicates(keep='last')
        
        # Clear the existing data in the sheet
        sheet.delete_rows(2, sheet.max_row)
        
        # Convert the updated DataFrame to rows
        rows = dataframe_to_rows(updated_df, index=False, header=False)
        
        # Write the headers if they are not already present
        for c_idx, header in enumerate(updated_df.columns, 1):
            sheet.cell(row=1, column=c_idx, value=header)
        
        # Write the updated rows to the sheet
        for r_idx, row in enumerate(rows, 2):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Save the workbook
        book.save(output_file)
        logger.info(f"DataFrame updated and saved as {output_file}")

def main():
    all_data = pd.DataFrame(columns=columns)
    
    for form_id in form_ids:
        full_api_url = f'{base_api_url}{form_id}'
        data = get_form_submissions(full_api_url, username, password)
        if data:
            form_submissions = data.get('form_submissions', [])
            for submission in form_submissions:
                processed_submission = process_submission(submission, form_id)
                all_data = pd.concat([all_data, pd.DataFrame([processed_submission])], ignore_index=True)
    
    output_file = 'form_submissions.xlsx'
    save_to_excel(all_data, output_file)

if __name__ == "__main__":
    main()